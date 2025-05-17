import os
import json
import time
import numpy as np
import re
import pdfplumber
from collections import Counter
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def save_checkpoint(all_transformed_questions, current_question_numberx, CHECKPOINT_FILE):
    """保存当前处理进度和结果到checkpoint文件"""
    try:
        checkpoint_data = {
            "all_transformed_questions": all_transformed_questions,
            "current_question_numberx": current_question_numberx
        }
        np.save(CHECKPOINT_FILE, checkpoint_data)
        print(f"检查点已保存到 {CHECKPOINT_FILE}。当前进度: {current_question_numberx}")

    except Exception as e:
        print(f"保存检查点时发生错误: {e}")

def load_checkpoint(CHECKPOINT_FILE):
    """加载之前的处理进度和结果"""
    try:
        if os.path.exists(CHECKPOINT_FILE):
            checkpoint = np.load(CHECKPOINT_FILE, allow_pickle=True).item()
            all_transformed = checkpoint.get("all_transformed_questions", [])
            current_question_numberx = checkpoint.get("current_question_numberx", 0)
            print(f"已加载检查点文件。已处理 {current_question_numberx} 个题目,已转换 {len(all_transformed)} 个结果。")
            return all_transformed, current_question_numberx
        else:
            print("未找到检查点文件,将从头开始处理。")
            return [], 0
    except Exception as e:
        print(f"加载检查点时发生错误: {e}")
        return [], 0

def extract_pdf_text(pdf_path, save_dir=None):
    """
    提取PDF文本并保存到文件中
    
    Args:
        pdf_path (str): PDF文件路径
        save_dir (str, optional): 保存目录，默认为PDF同目录
        
    Returns:
        str: 保存的文本文件路径
    """
    # 如果未提供保存目录，使用PDF所在目录
    if save_dir is None:
        save_dir = os.path.dirname(pdf_path)
    else:
        return
    
    # 创建保存目录（如果不存在）
    os.makedirs(save_dir, exist_ok=True)
    
    # 生成保存文件名（使用PDF文件名+.json）
    pdf_filename = os.path.basename(pdf_path)
    pdf_name_without_ext = os.path.splitext(pdf_filename)[0]
    save_path = os.path.join(save_dir, f"{pdf_name_without_ext}.json")
    
    # 检查是否已经提取过文本
    if os.path.exists(save_path):
        print(f"找到已提取的文本文件: {save_path}")
        return save_path
    
    # 开始提取
    print(f"开始提取PDF文本...")
    start_time = time.time()
    
    extracted_data = {
        "pdf_path": pdf_path,
        "extraction_time": time.strftime("%Y-%m-%d %H:%M:%S"),
        "total_pages": 0,
        "pages": {}
    }
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            extracted_data["total_pages"] = total_pages
                        
            for page_num, page in enumerate(pdf.pages):
                current_page = page_num + 1  # 页码从1开始
                print(f"正在提取第 {current_page}/{total_pages} 页...")
                
                # 提取文本
                text = page.extract_text() or ""
                extracted_data["pages"][str(current_page)] = text
        
        # 保存提取结果
        with open(save_path, "w", encoding="utf-8") as f:
            json.dump(extracted_data, f, ensure_ascii=False, indent=2)
        
        end_time = time.time()
        print(f"文本提取完成！用时 {end_time - start_time:.2f} 秒")
        print(f"文本已保存至: {save_path}")
    except:
        pass

def find_text_in_saved_pdf(text_file_path, search_text):
    """
    在已保存的PDF文本中查找指定文本
    
    Args:
        text_file_path (str): 保存的文本文件路径
        search_text (str): 需要查找的文本
    
    Returns:
        list: 包含文本出现页码的列表（页码从1开始）
    """
    result_pages = []
    
    try:
        # 读取保存的文本文件
        with open(text_file_path, "r", encoding="utf-8") as f:
            pdf_data = json.load(f)
        
        # 获取PDF总页数
        total_pages = pdf_data.get("total_pages", 0)
        pages_data = pdf_data.get("pages", {})
        
        print(f"PDF总页数: {total_pages}")
        
        # 遍历每一页
        for page_num in range(1, total_pages + 1):
            # 获取当前页面文本
            text = pages_data.get(str(page_num), "")
            
            # 检查搜索文本是否在当前页面
            if search_text in text:
                result_pages.append(page_num)
    
    except Exception as e:
        print(f"查找文本时发生错误: {e}")
    
    return result_pages

def find_mode(lst):
    """从列表中查找出现次数最多的元素"""
    count = Counter(lst)
    mode = count.most_common(1)[0][0]
    return mode

def is_chinese_char(char):
    """
    判断一个字符是否为常用的汉字。
    这里使用的是基本的CJK统一表意文字区段 (U+4E00 至 U+9FFF)。
    如果需要更广泛的汉字支持（例如扩展区A、B等），可以扩展此处的Unicode范围。
    """
    return "\u4e00" <= char <= "\u9fff"

def get_consecutive_chinese_chars(text):
    """
    从输入文本中随机查找一个由5个连续汉字组成的子字符串。
    """
    if not text or len(text) < 5:
        return ""  # 如果文本为空或长度小于10，则返回空字符串

    possible_substrings = []
    # 遍历所有可能的10字符子串的起始位置
    for i in range(len(text) - 4):  # 确保子字符串长度为10
        substring = text[i:i+5]
        # 检查子字符串中的所有字符是否都是汉字
        if all(is_chinese_char(char) for char in substring):
            possible_substrings.append(substring)
        
    return possible_substrings

def save_to_excel(data, file_path):
    """保存数据到Excel文件，包括图片处理"""
    # 清理非法字符
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")
    data = data.map(lambda x: ILLEGAL_CHARACTERS_RE.sub("", str(x)))

    # 获取图片基准目录（browser-llm目录）
    base_dir = os.path.abspath(os.path.join(os.path.dirname(file_path), '..'))

    try:
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        
        # 写入表头
        for col_idx, column in enumerate(data.columns, 1):
            ws.cell(row=1, column=col_idx, value=column)
        
        # 写入数据并处理图片
        for row_idx, row in enumerate(data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 检查是否是图片列（deepseek、千问、豆包）
                column_name = data.columns[col_idx-1]
                if column_name in ['deepseek', '千问', '豆包'] and value and isinstance(value, str) and value.endswith('.png'):
                    # 构建完整的图片路径
                    img_path = os.path.normpath(os.path.join(base_dir, value))
                    if os.path.exists(img_path):
                        try:
                            # 创建图片对象
                            img = Image(img_path)
                            # 调整图片大小
                            img.width = 20
                            img.height = 20
                            # 将图片添加到单元格
                            ws.add_image(img, cell.coordinate)
                            # 清空单元格中的图片路径文本
                            cell.value = ""
                        except Exception as e:
                            print(f"添加图片到 {cell.coordinate} 时出错: {str(e)}")
        
        # 保存工作簿
        wb.save(file_path)
        print(f"Excel文件已保存: {file_path}")
        
    except Exception as e:
        print(f"写入文件时出错: {e}") 